import numpy as np
import pandas as pd
import os
from torch.utils.tensorboard import SummaryWriter
import torch.cuda
from sklearn.model_selection import train_test_split
from torch.utils.data import DataLoader
import matplotlib.pyplot as plt
import openpyxl
from PhonemeDataset import *
from Model import *

device=torch.device("cuda" if torch.cuda.is_available() else "cpu")


hyper_param={
    'train_ratio':0.8,
    'random_state':20,
    'batch_size':1024,
    'epoches':500,
    'learning_rate':0.001,
    'correlation_threshold':0.2,
    'model_name':"PhonemePredict",
    'save_csv_path':"Submission"
}
def Load_Npy(npy_path,IsLabel):
    load_file=np.load(npy_path)
    label_list=[]
    if IsLabel:
        df=pd.DataFrame(load_file,dtype='int')

    else:
        df=pd.DataFrame(load_file,dtype='float')
    print("Load Done")
    return df
def Train_Load_Optimize(input_dim):
    model=Model(input_dim,device)
    model=model.to(device)
    model.loss=model.loss.to(device)
    optim=torch.optim.Adam(model.parameters(),hyper_param['learning_rate'])
    step=0
    pre_total_loss=999999999999
    loss1=nn.CrossEntropyLoss()
    loss1=loss1.to(device)
    for ecpoch in range(hyper_param['epoches']):
        model.train()
        total_train_loss=0.0
        total_valid_loss=0.0
        total_accuracy=0
        for data in train_dataloader:
            inputs,targets=data
            inputs,targets=inputs.to(device),targets.to(device)
            targets=targets.squeeze(1)
            outputs=model(inputs)
            loss=model.cal_loss(outputs,targets)
            optim.zero_grad()
            loss.backward()
            optim.step()
            total_train_loss+=loss
        model.eval()
        with torch.no_grad():
            for data in valid_dataloader:
                inputs,targets=data
                inputs,targets=inputs.to(device),targets.to(device)
                targets = targets.squeeze(1)
                outputs=model(inputs)
                loss=model.cal_loss(outputs,targets)
                total_valid_loss+=loss
                accuracy=(outputs.argmax(1)==targets).sum()/hyper_param['batch_size']
                total_accuracy+=accuracy
        step+=1

        if (step is 1) or pre_total_loss>(total_valid_loss+total_train_loss):
            torch.save(model.state_dict(),"C:/Users/Jian/Desktop/"+hyper_param['model_name']+".pth")
            pre_total_loss=total_valid_loss+total_train_loss
            print("*{}".format(step) + "TrainLoss: {}".format(total_train_loss) + "   ValidLoss: {}".format(
                total_valid_loss)+"   AccuracyRate: {}".format(total_accuracy/hyper_param['batch_size']))
        else:
            print("{}".format(step) + "TrainLoss: {}".format(total_train_loss) + "   ValidLoss: {}".format(
                total_valid_loss)+"   AccuracyRate: {}".format(total_accuracy/hyper_param['batch_size']))


def Test(input_dim,load_model_name):
    model=Model(input_dim,device)
    model.load_state_dict(torch.load(load_model_name+".pth"))
    pred_list=[]
    for data in test_dataloader:
        preds=model(data)
        for pred in preds.argmax(1).numpy().tolist():
            pred_list.append(pred)
    work_book=openpyxl.Workbook()
    work_sheet=work_book.create_sheet('Submission')
    row=1
    column=1
    for id,pred_class in enumerate(pred_list):
        work_sheet.cell(row,column).value=id
        column+=1
        work_sheet.cell(row,column).value=pred_class
        row+=1
        column=1
    del work_book["Sheet"]
    work_book.save(hyper_param['save_csv_path']+'.csv')
def Cal_Correlation():
    remove_col_lsit=[]
    result_series=train_label_df[0]
    for col_index in range(429):
        feature_df=train_df.iloc[:,[col_index]]
        feature_series=feature_df[feature_df.columns[0]]
        corr=np.corrcoef(feature_series,result_series)[0,1]
        if abs(corr)<hyper_param['correlation_threshold']:
            remove_col_lsit.append(col_index)
    train_df2=train_df.drop(train_df.iloc[:,remove_col_lsit],axis=1)
    test_df2=test_df.drop(train_df.iloc[:,remove_col_lsit],axis=1)
    return train_df2,test_df2

train_df=Load_Npy("C:/Users/Jian/Desktop/train_11.npy",IsLabel=False)
train_label_df=Load_Npy("C:/Users/Jian/Desktop/train_label_11.npy",IsLabel=True)
test_df=Load_Npy("C:/Users/Jian/Desktop/test_11.npy",IsLabel=False)
# train_df,test_df=Cal_Correlation()
x_train,x_valid,y_train,y_valid=train_test_split(train_df,train_label_df,train_size=hyper_param['train_ratio'],random_state=hyper_param['random_state'])
test_dataset=PhonemeDataset(test_df)
valid_dataset=PhonemeDataset(x_valid,y_valid)
train_dataset=PhonemeDataset(x_train,y_train)

train_dataloader=DataLoader(train_dataset,hyper_param['batch_size'],shuffle=True)
valid_dataloader=DataLoader(valid_dataset,hyper_param['batch_size'],shuffle=True)
test_dataloader=DataLoader(test_dataset,hyper_param['batch_size'],shuffle=False)

print("DataLoader Done")

#Train_Load_Optimize(train_df.shape[1])
Test(train_df.shape[1],"C:/Users/Jian/Desktop/"+hyper_param['model_name'])

